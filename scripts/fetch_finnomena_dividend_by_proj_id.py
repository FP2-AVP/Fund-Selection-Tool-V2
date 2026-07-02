#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import ssl
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

ROOT = Path(__file__).resolve().parents[1]
SEC_API_PATH = ROOT / "Data" / "Data For SEC API - 2026-Q1.json"
FUND_KEY_PERFORMANCE_PATH = ROOT / "Data" / "Fund Key Performance AVP - 2026-Q1.json"
MASTER_PROFILES_PATH = ROOT / "Python By Boss เพื่อดึงข้อมูล" / "fund_master_profiles.xlsx"


def load_row_map(path: Path, key_name: str) -> dict[str, dict[str, Any]]:
    rows = json.loads(path.read_text())
    headers = rows[0]
    result: dict[str, dict[str, Any]] = {}
    try:
        key_idx = headers.index(key_name)
    except ValueError as exc:
        raise SystemExit(f"Missing column {key_name!r} in {path}") from exc

    for row in rows[1:]:
        if key_idx >= len(row):
            continue
        key = str(row[key_idx]).strip().upper()
        if not key:
            continue
        result[key] = dict(zip(headers, row))
    return result


def load_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    rows = json.loads(path.read_text())
    return rows[0], rows[1:]


def load_master_profile_records() -> list[dict[str, Any]]:
    if load_workbook is None or not MASTER_PROFILES_PATH.exists():
        raw_headers, raw_rows = load_rows(SEC_API_PATH)
        return [dict(zip(raw_headers, row)) for row in raw_rows]

    wb = load_workbook(MASTER_PROFILES_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v) if v is not None else "" for v in next(rows)]
    records: list[dict[str, Any]] = []
    for row in rows:
        record = dict(zip(headers, row))
        records.append(
            {
                "Proj_id": record.get("proj_id"),
                "proj_abbr_name": record.get("proj_abbr_name"),
                "fund_class_name": record.get("fund_class_name"),
                "Fund Name": record.get("proj_name_en") or record.get("proj_name_th") or record.get("proj_abbr_name"),
                "fund_status": record.get("fund_status"),
                "comp_name_th": record.get("comp_name_th"),
                "comp_name_en": record.get("comp_name_en"),
                "unique_id": record.get("unique_id"),
                "fund_class_detail": record.get("fund_class_detail"),
                "last_upd_date": record.get("last_upd_date"),
            }
        )
    return records


def build_proj_id_mapping() -> dict[str, list[dict[str, Any]]]:
    perf_map = load_row_map(FUND_KEY_PERFORMANCE_PATH, "Fund Code")
    result: dict[str, list[dict[str, Any]]] = {}

    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in load_master_profile_records():
        proj_id = str(record.get("Proj_id") or "").strip().upper()
        if not proj_id:
            continue
        grouped.setdefault(proj_id, []).append(record)

    for proj_id, sec_rows in grouped.items():
        candidates: list[tuple[str, dict[str, Any], str]] = []
        seen_codes: set[str] = set()
        for sec_row in sec_rows:
            for field_name, match_source in (("fund_class_name", "fund_class_name"), ("proj_abbr_name", "proj_abbr_name")):
                fund_code = str(sec_row.get(field_name) or "").strip().upper()
                if not fund_code or fund_code in seen_codes:
                    continue
                perf_row = perf_map.get(fund_code)
                if not perf_row:
                    continue
                fund_id = str(perf_row.get("SecId") or "").strip()
                if not fund_id:
                    continue
                seen_codes.add(fund_code)
                candidates.append((fund_code, sec_row, match_source))

        if not candidates:
            continue

        result[proj_id] = []
        for fund_code, sec_row, match_source in candidates:
            perf_row = perf_map[fund_code]
            result[proj_id].append(
                {
                    "proj_id": proj_id,
                    "fund_code": fund_code,
                    "fund_name": perf_row.get("Name") or sec_row.get("Fund Name") or fund_code,
                    "asset_house": perf_row.get("Asset House") or "",
                    "dividend_label": perf_row.get("Dividend") or "",
                    "finnomena_fund_id": str(perf_row.get("SecId") or "").strip(),
                    "sec_proj_abbr_name": str(sec_row.get("proj_abbr_name") or "").strip(),
                    "sec_fund_class_name": str(sec_row.get("fund_class_name") or "").strip(),
                    "sec_fund_status": str(sec_row.get("fund_status") or "").strip(),
                    "sec_unique_id": str(sec_row.get("unique_id") or "").strip(),
                    "sec_row": sec_row,
                    "percentrank_row": perf_row,
                    "match_source": match_source,
                }
            )
    return result


def resolve_proj_id(proj_id: str) -> list[dict[str, Any]]:
    mapping = build_proj_id_mapping()
    sec_row = mapping.get(proj_id.upper())
    if not sec_row:
        raise SystemExit(f"proj_id not found in {SEC_API_PATH.name}: {proj_id}")
    return sec_row


def fetch_finnomena_dividend_history(fund_id: str) -> dict[str, Any]:
    url = f"https://www.finnomena.com/fn3/api/fund/v2/public/funds/{fund_id}/dividend"
    req = Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0",
        },
    )
    try:
        with urlopen(req, timeout=45, context=ssl._create_unverified_context()) as resp:
            body = resp.read().decode("utf-8")
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} from Finnomena for fund_id {fund_id}: {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"Network error while fetching Finnomena data for fund_id {fund_id}: {exc}") from exc

    try:
        return json.loads(body)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid JSON from Finnomena for fund_id {fund_id}") from exc


def build_result(resolved: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    dividends = ((payload or {}).get("data") or {}).get("dividends") or []
    return {
        "proj_id": resolved["proj_id"],
        "fund_code": resolved["fund_code"],
        "match_source": resolved.get("match_source", ""),
        "fund_name": resolved["fund_name"],
        "asset_house": resolved["asset_house"],
        "dividend_label": resolved["dividend_label"],
        "finnomena_fund_id": resolved["finnomena_fund_id"],
        "sec_proj_abbr_name": resolved.get("sec_proj_abbr_name", ""),
        "sec_fund_class_name": resolved.get("sec_fund_class_name", ""),
        "sec_fund_status": resolved.get("sec_fund_status", ""),
        "sec_unique_id": resolved.get("sec_unique_id", ""),
        "fetch_error": "",
        "finnomena_status": payload.get("status") if isinstance(payload, dict) else None,
        "dividend_count": len(dividends),
        "dividends": dividends,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Fetch Finnomena dividend history using SEC proj_id as the lookup key.")
    parser.add_argument("proj_id", nargs="?", help="SEC proj_id เช่น M0034_2539")
    parser.add_argument("--all-matched", action="store_true", help="Fetch every proj_id that can be mapped from local SEC data to Finnomena fund id")
    parser.add_argument("--limit", type=int, default=0, help="Limit rows when using --all-matched")
    parser.add_argument("--workers", type=int, default=12, help="Parallel workers when using --all-matched")
    parser.add_argument("--output", help="Write JSON output to this file path")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON output")
    args = parser.parse_args()

    if args.all_matched:
        mapping = build_proj_id_mapping()
        items = [item for group in mapping.values() for item in group]
        if args.limit > 0:
            items = items[: args.limit]
        output = []
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
            future_map = {
                executor.submit(fetch_finnomena_dividend_history, resolved["finnomena_fund_id"]): resolved
                for resolved in items
            }
            for future in as_completed(future_map):
                resolved = future_map[future]
                try:
                    payload = future.result()
                    output.append(build_result(resolved, payload))
                except Exception as exc:
                    output.append(
                        {
                            "proj_id": resolved["proj_id"],
                            "fund_code": resolved["fund_code"],
                            "fund_name": resolved["fund_name"],
                            "asset_house": resolved["asset_house"],
                            "dividend_label": resolved["dividend_label"],
                            "finnomena_fund_id": resolved["finnomena_fund_id"],
                            "sec_proj_abbr_name": resolved.get("sec_proj_abbr_name", ""),
                            "sec_fund_class_name": resolved.get("sec_fund_class_name", ""),
                            "sec_fund_status": resolved.get("sec_fund_status", ""),
                            "sec_unique_id": resolved.get("sec_unique_id", ""),
                            "fetch_error": str(exc),
                            "finnomena_status": False,
                            "dividend_count": 0,
                            "dividends": [],
                        }
                    )
        output.sort(key=lambda item: item["proj_id"])
        result: Any = {
            "matched_proj_id_count": len(mapping),
            "matched_count": len(items),
            "fetched_count": len(output),
            "items": output,
        }
    else:
        if not args.proj_id:
            raise SystemExit("proj_id is required unless --all-matched is used")
        resolved_list = resolve_proj_id(args.proj_id)
        output = []
        for resolved in resolved_list:
            try:
                payload = fetch_finnomena_dividend_history(resolved["finnomena_fund_id"])
                output.append(build_result(resolved, payload))
            except Exception as exc:
                output.append(
                    {
                        "proj_id": resolved["proj_id"],
                        "fund_code": resolved["fund_code"],
                        "match_source": resolved.get("match_source", ""),
                        "fund_name": resolved["fund_name"],
                        "asset_house": resolved["asset_house"],
                        "dividend_label": resolved["dividend_label"],
                        "finnomena_fund_id": resolved["finnomena_fund_id"],
                        "sec_proj_abbr_name": resolved.get("sec_proj_abbr_name", ""),
                        "sec_fund_class_name": resolved.get("sec_fund_class_name", ""),
                        "sec_fund_status": resolved.get("sec_fund_status", ""),
                        "sec_unique_id": resolved.get("sec_unique_id", ""),
                        "fetch_error": str(exc),
                        "finnomena_status": False,
                        "dividend_count": 0,
                        "dividends": [],
                    }
                )
        result = {
            "proj_id": args.proj_id.upper(),
            "matched_count": len(output),
            "items": output,
        }

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2 if args.pretty else None) + "\n")
    else:
        json.dump(result, sys.stdout, ensure_ascii=False, indent=2 if args.pretty else None)
        sys.stdout.write("\n")


if __name__ == "__main__":
    main()
